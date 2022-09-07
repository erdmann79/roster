#!bin/env/python

"""This module does blah blah."""

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pprint import pprint

import pandas


class Student(object):
    def __init__(self, name, id_number, student_sheet):
        """
        Init SampleClass with blah.

        Args:
            filename: Name of the workbook file.
        """
        self.name = name
        self.first_name = name.split(" ")[0]
        self.last_name = name.split(" ")[1]
        self.id = id_number
        self.student_sheet_title = student_sheet.title
        self.student_sheet = student_sheet
        data = student_sheet.values
        next(data)
        next(data)
        next(data)
        next(data)
        columns = next(data)[0:]
        dataframe = pandas.DataFrame(data, columns=columns)
        dataframe = dataframe.rename(columns={"Grade": "grades"})
        self.grades = dataframe["grades"]

    def __str__(self):
        """Return item name."""
        return self.name

    def __getitem__(self, item):
        """Get item by index."""
        return getattr(self, item)

    def __setitem__(self, key, value):
        """Set item by index."""
        setattr(self, key, value)


class Roster(object):
    """
    Summery of Roster Class.

    Mrs. Jones, a teacher at Billings Elementary School, has been asked by the administration to
    track her students' grades in a roster stored in an Excel file. Unfortunatly, Mrs. Jones had a
    bad experience with Excel as a child and as solemnly sworn never to use Excel again. Your task
    is to create a Python class for reading and manipulating a provided Excel file so Mrs. Jones
    doesn't haven't to use Excel to successfully record her students' grades.

    """

    def __init__(self, filename):
        """
        Init SampleClass with blah.

        Args:
            filename: Name of the workbook file.
        """
        # for student_name in get_student_names():
        #    print(student_name['firstname'])

        self.filename = filename
        self.students = []

    def __enter__(self):
        """Open the workbook."""
        self.workbook = load_workbook(self.filename)
        self.roster_sheet = self.workbook["Roster"]
        for name in self.get_student_names():
            id_number = self.get_id(name)
            student = Student(name, id_number, self.get_student_sheet(name))
            self.students.append(student)
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        """Close the workbook."""
        self.workbook.close()

    def get_id(self, name):
        """
        Get student id.

        Returns:
            int student id

        """
        # roster_sheet = self.workbook["Roster"]
        id_number = int()
        for (id_number, first, last) in self.roster_sheet.iter_rows(
            min_row=2, min_col=1, max_col=3, values_only=True
        ):
            if not first and not last:
                break
            if name == first + " " + last:
                break
        return id_number

    def get_student_sheet(self, name):
        """
        Get student sheet.

        Returns:
            openpyxl sheet of student

        """
        # roster_sheet = self.workbook["Roster"]
        student_sheet = None
        for i, student_sheet in enumerate(self.workbook):
            if i == 0:
                continue
            if name == student_sheet["B2"].internal_value:
                break
        return student_sheet

    def get_student_names(self):
        """
        Get student names.

        Returns:
            List of student names

        """
        # roster_sheet = self.workbook["Roster"]

        student_name_list = []
        for (first, last) in self.roster_sheet.iter_rows(
            min_row=2, min_col=2, max_col=3, max_row=1000, values_only=True
        ):
            if not first and not last:
                break
            student_name_list.append(first + " " + last)

        return student_name_list

    def get_student(self, name):
        """
        Get student.

        Args:
            name: Name of the student.

        Returns:
            This is a description of what is returned.

        """
        student_obj = None
        for student_obj in self.students:
            if name == student_obj["name"]:
                break
        return student_obj

    def class_average(self):
        """
        Get class average.

        Returns:
            int class average.

        #    self.assertTrue(roster_obj.class_average() == 614.1 / 7)

        """
        student_avg_list = []
        for student in self.students:
            student_avg_list.append(student["grades"].mean())
        class_average = sum(student_avg_list) / len(student_avg_list)
        return class_average

    def delete_student(self, name):
        """
        Remove a student from the workbook.

        Args:
            filename: Name of the student.
        """
        student_obj = self.get_student(name)
        self.workbook.remove(student_obj["student_sheet"])
        student_list = self.get_student_names()
        shift_register = 2
        for i, student_name in enumerate(student_list):
            student_obj = self.get_student(student_name)
            row = i + shift_register
            id_number = i + shift_register - 1
            if student_name == name:
                # print("remove", name, "row", row, "id", id_number)
                self.roster_sheet.delete_rows(row)
                shift_register = 1
                continue
            if shift_register == 1:
                self.roster_sheet[f"A{row}"] = id_number
                student_obj["student_sheet"]["B1"] = id_number
                student_obj["id"] = id_number
                student_obj["student_sheet_title"] = f"Student_{id_number}"
                student_obj["student_sheet"].title = f"Student_{id_number}"

    def save(self, filename=None):
        """
        Write workbook to file.

        Args:
            filename: Name of the workbook file.

        """
        row_offset = 3
        for student in self.students:
            dataframe = pandas.DataFrame(student["grades"])
            rows = dataframe_to_rows(dataframe)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    if r_idx == 1 or r_idx == 2:
                        continue
                    student["student_sheet"].cell(
                        row=r_idx + row_offset, column=c_idx, value=value
                    )

        if not filename:
            filename = self.filename

        self.workbook.save(filename)


if __name__ == "__main__":
    with Roster("Jones_2019.xlsx") as roster_obj:
        print("***** Main ************************")

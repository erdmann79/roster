import unittest
import pandas
from openpyxl import load_workbook
from roster import Roster


class TestRoster(unittest.TestCase):
    """Roster slass tests."""

    def test_read_roster(self):
        """Test read roster_obj."""
        with Roster("Jones_2019.xlsx") as roster_obj:
            student_names = roster_obj.get_student_names()
            self.assertTrue(len(student_names) == 7)
            self.assertTrue("Robert Waters" in student_names)

            catherine = roster_obj.get_student("Catherine Hitchens")
            self.assertTrue(catherine["id"] == 3)
            self.assertTrue(isinstance(catherine["grades"], pandas.Series))
            self.assertTrue(len(catherine["grades"]) == 10)
            self.assertTrue(catherine["grades"][4] == 86)

            self.assertTrue(roster_obj.class_average() == 614.1 / 7)

    def test_write_roster(self):
        """Test write roster_obj."""
        with Roster("Jones_2019.xlsx") as roster_obj:
            john = roster_obj.get_student("Johnny Carson")
            for assignment, grade in [(3, 90), (6, 94), (9, 92)]:
                john["grades"][assignment] = grade
            self.assertTrue(roster_obj.class_average() == 616.6 / 7)
            roster_obj.save("Jones_2019_Updated.xlsx")

        workbook = load_workbook("Jones_2019_Updated.xlsx")
        self.assertTrue(workbook.get_sheet_by_name("Student_1")["B12"].value == 94)
        workbook.close()

    def test_delete_roster_student(self):
        """Test delete roster student."""
        student_count = 0
        with Roster("Jones_2019.xlsx") as roster_obj:
            student_count = len(roster_obj.get_student_names())
            self.assertTrue(student_count == 7)
            self.assertTrue(roster_obj.get_student("William Thomas")["id"] == 5)
            roster_obj.delete_student("Allen Dalton")
            student_count = len(roster_obj.get_student_names())
            self.assertTrue(student_count == 6)
            self.assertTrue(roster_obj.get_student("William Thomas")["id"] == 4)
            roster_obj.save("Jones_2019_Reduced.xlsx")

        workbook = load_workbook("Jones_2019_Reduced.xlsx")
        sheet_names = workbook.sheetnames
        self.assertTrue(len(sheet_names) == 7)
        self.assertTrue(sheet_names[0] == "Roster")
        self.assertTrue(sheet_names[-1] == "Student_6")
        self.assertTrue(workbook["Student_3"]["B7"].value == 92)
        workbook.close()


if __name__ == "__main__":
    unittest.main()
